import praw
import prawcore
import re
from datetime import datetime
from openpyxl import load_workbook
import sys

r = praw.Reddit(client_id = 'YOUR_CLIENT_ID_HERE', client_secret = 'YOUR_CLIENT_SECRET_HERE', user_agent = 'YOUR_USER_AGENT_HERE')
print("Logging on...")

namefile = open('countryname.txt','r')
#alias_dict = {}
count_dict = {}
country_list = []

wb = load_workbook(filename='polandball_main.xlsx')
ws = wb['INPUT']
if ws.cell(row=1,column=3).value != "id":
	print ("ID cannot be found")
	sys.exit()
	
alias_dict = ({},{},{})
special = ['@','%','<','>']
len_sp = len(special)-1
len_al = len(alias_dict)-1

for line in namefile:
	pair = line.rstrip().split('=',1)
	if len(pair) == 1:
		pair.insert(0,pair[-1].split('/')[0])
	for i in pair[-1].split('/'):
		if i[0] == '+':
			alias_dict[0][i[1:]] = pair[0]
		elif i[0] == '-':
			alias_dict[2][i[1:]] = pair[0]
		else:
			alias_dict[1][i] = pair[0]
		#alias_dict[i] = pair[0]
	country_list.append(pair[0])

def dropchar(text):
	for c in special:
		text = text.replace(c,'')
	return re.sub('[^a-zA-Z ]', ' ', text)
	
def match_country(match):
	if match in parsed:
		if parsed.count(match) > 1:
			print (logtext + ' ' + alias + ' x' + str(parsed.count(match)))
		else:
			print (logtext + ' ' + alias)
		count_dict[alias_dict[alias]] += multiplier
		parsed = parsed.replace(match,'')

def find_country():
	lpad = '' if '<' in alias[0:len_sp] else ' '
	rpad = ' ' if '>' in alias[0:len_sp] else ''
	content = dropchar(alias)
	match_country(lpad + content + rpad)
	
	if '%' in alias[0:len_sp]:
		content = '.'.join(e for e in alias if e.isalpha())
		match_country(lpad + content + rpad)	
						
j = 10000#2
for row in range(j,len(ws['C'])):
	post = r.submission(id=ws.cell(row=j,column=3).value)
#for post in r.subreddit('polandball').new(limit=None):
	post.comments.replace_more(limit=None)
	
	print ("###########################################################")
	count_dict = dict.fromkeys(country_list,0) #reset count
	
	for comment in post.comments.list():
		parsed = ' ' + dropchar(comment.body) + ' ' #padding
		print (parsed)
		
		#OP bonus
		if post.author == comment.author:
			multiplier = 10
			logtext = "     [OP]"
		else:
			multiplier = 1
			logtext = "     [User]"
		
		#first pass @
		for priority in range(0,len_al):
			for alias in alias_dict[priority]:	
				if '@' in alias[0:len_sp]:
					find_country()
		#lower pass
		parsed = parsed.lower()
		for priority in range(0,len(alias_dict)-1):
			for alias in alias_dict[priority]:
				if '@' not in alias[0:len_sp]:
					find_country()
	#title
	parsed = ' ' + dropchar(post.title) + ' '
	multiplier = 20
	logtext = "     [TITLE]"

	for priority in range(0,len_al):
		for alias in alias_dict[priority]:	
			if '@' in alias[0:len_sp]:
				find_country()
	parsed = parsed.lower()
	for priority in range(0,len_al):
		for alias in alias_dict[priority]:
			if '@' not in alias[0:len_sp]:
				find_country()
				
	#count stats
	total = 0.0
	output = []
	out = ''
	
	for country in count_dict:
		if count_dict[country] != 0:
			total += count_dict[country]
			
	for country in count_dict:
		if count_dict[country] != 0:
			output.append([count_dict[country],count_dict[country]/total,country])
	output.sort(reverse=True)
	for i in output:
		print ("[" + str(i[0]) + "] " + i[2] + ": " + str(i[1]))
		if i[1] >= 0.05:
			out += i[2] + ":" + str(round(i[1],2)) + "/"
	
	print(str(datetime.utcfromtimestamp(post.created_utc)) + ": " + post.title.encode('utf-8'))
	print(post.url.encode('utf-8'))
	ws.cell(row=j,column=12).value = out[:-1]
	print ws.cell(row=j,column=12).value
	j += 1
	
wb.save('test_cname.xlsx')