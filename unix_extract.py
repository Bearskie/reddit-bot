"""
Extracts all posts from a subreddit, outputs data in excel format.

PRAW Unix Extract
Made for r/polandball by u/Barskie
"""

import praw
import prawcore
import time
import sys
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook

r = praw.Reddit(client_id = 'YOUR_CLIENT_ID_HERE', client_secret = 'YOUR_CLIENT_SECRET_HERE', user_agent = "YOUR_USER_AGENT_HERE")
print("Logging on...")

sub_name = 'polandball'
sub = r.subreddit(sub_name)
file = sub_name + '_main.xlsx'
sheet = 'INPUT'
rest = 604800 #1 week
clock = time.time() - rest #only fetch posts older than 1 week, give time for scores to stabilize

if os.path.isfile(file) == True:
	wb = load_workbook(filename=file)
	ws = wb[sheet]
	wb.save(os.path.splitext(file)[0] + '_backup.xlsx')
else:
	wb = Workbook()
	ws = wb.active
	ws.title = sheet
	ws.cell(row = 1, column = 1).value = "Timestamp"
	ws.cell(row = 1, column = 2).value = "Date"
	ws.cell(row = 1, column = 3).value = "id"
	ws.cell(row = 1, column = 4).value = "Flair (Post)"
	ws.cell(row = 1, column = 5).value = "Flair (User)"
	ws.cell(row = 1, column = 6).value = "User"
	ws.cell(row = 1, column = 7).value = "Title"
	ws.cell(row = 1, column = 8).value = "Score"
	ws.cell(row = 1, column = 9).value = "Permalink"
	ws.cell(row = 1, column = 10).value = "Comments"
	ws.cell(row = 1, column = 11).value = "Directlink"
	print ("File '" + str(file) + "' does not exist. Creating new file...")
	
update_count = 0
write_count = 0
id_dict = {}
flair_dict = {}
utc_list = []
size = ws.max_row

for row in ws.iter_rows(min_row=1, max_col=3, max_row=size):
	if row[0].value == None:
		size = size-1
		
i = size + 1 #row writer, start write from which row
j = 2 #row tracker, start read from which row, later used as update writer

if j > size:
	archive_point = sub.created_utc
else:
	for row in ws.iter_rows(min_row=j, max_col=6, max_row=size):
		id_dict[row[2].value] = j
		flair_dict[row[5].value] = row[4].value
		utc_list.append(row[0].value)
		j=j+1
	archive_point = max(utc_list) - 15552000 #180 days, posts are archived

#functions
def write_excel():
	ws.cell(row = i, column = 1).value = post.created_utc
	ws.cell(row = i, column = 2).value = datetime.utcfromtimestamp(post.created_utc)
	ws.cell(row = i, column = 2).number_format = 'D/M/YYYY HH:MM:SS'
	ws.cell(row = i, column = 3).value = str(post.id.encode('utf-8'))
	ws.cell(row = i, column = 4).value = str(post.link_flair_css_class)
	if post.author_flair_css_class != None:
		flair_dict[post.author.name] = post.author_flair_css_class
		ws.cell(row = i, column = 5).value = str(post.author_flair_css_class)
	ws.cell(row = i, column = 6).value = str(post.author.name)
	ws.cell(row = i, column = 7).value = str(post.title.encode('utf-8'))
	ws.cell(row = i, column = 8).value = post.score
	ws.cell(row = i, column = 9).value = "https://redd.it/" + str(post.id.encode('utf-8')) #str("www.reddit.com" + post.permalink.encode('utf-8'))
	ws.cell(row = i, column = 10).value = post.num_comments
	ws.cell(row = i, column = 11).value = str(post.url.encode('utf-8'))
	
def update_excel():
	if post.author_flair_css_class != None:
		flair_dict[post.author.name] = post.author_flair_css_class
		ws.cell(row = j, column = 5).value = str(post.author_flair_css_class)
	ws.cell(row = j, column = 8).value = post.score
	ws.cell(row = j, column = 10).value = post.num_comments
	
def update_flair():
	delta = []
	for k in range(2,size):
		user = ws.cell(row = k, column = 6).value
		flair = ws.cell(row = k, column = 5).value
		if flair != flair_dict[user]:
			if user not in delta:
				print (str(user) + ": " + flair + " --> " + flair_dict[user])
				delta.append(user)
			ws.cell(row = k, column = 5).value = flair_dict[user]
			
#time is always ROUNDED DOWN ie 1000.9 > 1000, so use integers between ranges
while clock >= archive_point:
	try:
		for post in sub.submissions(start = max(clock - rest + 1, archive_point), end = clock):
			if post.id in id_dict:
				j = id_dict[post.id]
				update_excel()
				update_count = update_count+1
				print ("UPDATE: " + str(datetime.utcfromtimestamp(post.created_utc)) + " " + str(post.author.name) + " - " + post.title.encode('utf-8'))
			else:
				write_excel()
				write_count = write_count+1
				i = i+1
				print ("WRITE: " + str(datetime.utcfromtimestamp(post.created_utc)) + " " + str(post.author.name) + " - " + post.title.encode('utf-8'))
		clock = clock - rest
		
	except prawcore.exceptions.ServerError:
		print("Resting... (server error)")
		time.sleep(30)
		print(post.created_utc)
		clock = post.created_utc
	
	except prawcore.exceptions.RequestException:
		print("Resting... (request error)")
		time.sleep(30)
		print(post.created_utc)
		clock = post.created_utc
		
	except KeyboardInterrupt:
		print(post.created_utc)
		wb.close()
		sys.exit()

update_flair()
wb.save(file)
print ("Initial number of entries: " + str(size))
print (str(write_count) + " written. " + str(update_count) + " updated.")
print ("Current number of entries: " + str(size + write_count))
